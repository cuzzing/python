#!/usr/bin/env python
# encoding=utf-8

__author__ = 'steady_animal'

from openpyxl import load_workbook
import time

fileName = 'pythonExcel.xlsx'
wb = load_workbook(fileName)
print(wb.get_sheet_names())
ws = wb.active

# 两种写入数据的方法
# ws['A3'] = 111
# ws.cell(row=4, column=3, value=10)

row_length = len(list(ws.rows))  # 获取行数
column_length = len(list(ws.columns))  # 获取列数
# print(row_length)
time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
ws.cell(row=1, column=column_length + 1, value=time)

for i in range(2, row_length):
    ws.cell(row=i, column=column_length + 1, value=None)

wb.save(fileName)
